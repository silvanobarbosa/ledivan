#!/usr/bin/env python3
"""
Script COMPLETO v2 de importação dos dados da Gisele
Planilha: Financeiro 040826.xlsx

Importa:
- Pacientes (da aba "Financ Cobranças")
- Metadata dos pacientes (dia atendimento, horário, frequência, etc.)
- Sessões/Agendas com presença e cobrança (da aba "Financ Cobranças")
- Pagamentos (da aba "Financ entradas")
"""

import openpyxl
import psycopg2
from datetime import datetime, date
import uuid
import os
from decimal import Decimal
import re

# Configuração
TENANT_ID = '5588fb64-368f-4deb-81be-80a611967ec1'  # Gisele
DATABASE_URL = os.getenv('DATABASE_URL')

if not DATABASE_URL:
    print("Erro: DATABASE_URL não definida")
    exit(1)

# Conectar ao banco
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

# Nomes que NÃO são pacientes (despesas/outros)
NAO_PACIENTES = {
    'Academia', 'Botox', 'Bolo Gi', 'Capoeira', 'Cartão', 'Ceclin',
    'Congresso', 'Custo fixo', 'Custo Fixo', 'Dentista Heitor', 'Depilação',
    'Devolução Amazon', 'Diferença', 'Diferença:', 'Doação Obreiros',
    'Débito', 'Entrada Diversos', 'Entrada Pontos Cartão', 'Entradas div',
    'Entrada Heitor', 'Estacionamento', 'Exame', 'Festa Vini', 'Heitor escola',
    'Heitor Escola', 'IPTU', 'Iptu', 'Investimento', 'Manutenção',
    'Matricula Heitor', 'Mercado', 'Multas', 'Pagamentos', 'Pedicure',
    'Pix enviado diversos', 'Pix Enviado Diversos', 'Placa dentista',
    'Placa Dentista', 'Present Cris', 'Present EU', 'Present Eu',
    'Present Silvano', 'Presente Felipe', 'Presente Laura', 'Presente Paty',
    'Presente Rafa', 'Pós', 'RNTP', 'Rntp', 'Remédios', 'Sala',
    'Supervisão', 'Terapia', 'Terapia Heitor', 'Uber', 'Vereda'
}

def normalizar_nome(nome):
    """Normaliza nome: primeira letra maiúscula, resto minúsculo"""
    if not nome:
        return ''
    return ' '.join(p.capitalize() for p in str(nome).strip().split())

def is_paciente(nome):
    """Verifica se é nome de paciente (não despesa)"""
    if not nome or not isinstance(nome, str):
        return False

    norm = normalizar_nome(nome)

    if nome in NAO_PACIENTES or norm in NAO_PACIENTES:
        return False

    # Se tem apenas letras e espaços, provavelmente é paciente
    return all(c.isalpha() or c.isspace() for c in norm) and len(norm) > 2

def excel_date_to_python(excel_date):
    """Converte data do Excel para Python date"""
    if not excel_date:
        return None

    if isinstance(excel_date, datetime):
        return excel_date.date()

    if isinstance(excel_date, date):
        return excel_date

    if isinstance(excel_date, str):
        # Try parsing DD/MM/YYYY format
        match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', excel_date)
        if match:
            day, month, year = map(int, match.groups())
            return date(year, month, day)
        return None

    if isinstance(excel_date, (int, float)):
        # Excel usa 1900-01-01 como base (serial 1)
        # Mas tem bug: considera 1900 bissexto (não é)
        from datetime import timedelta
        base = datetime(1899, 12, 30)  # Ajuste do bug
        try:
            return (base + timedelta(days=excel_date)).date()
        except:
            return None

    return None

def extract_valor_sessao(texto):
    """Extrai valor numérico de texto como '80,00 - reajustado' ou '50'"""
    if not texto:
        return 200  # Default

    texto = str(texto).replace(',', '.')
    match = re.search(r'(\d+(?:\.\d+)?)', texto)
    if match:
        return float(match.group(1))
    return 200

print('Iniciando importação COMPLETA v2 dos dados da Gisele...\n')

# Carregar planilha
file_path = 'C:/Users/User/Ledivan Plus/Gi/Financeiro 040826.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Verificar usuário
cur.execute('SELECT id, name, email FROM "user" WHERE id = %s', (TENANT_ID,))
user = cur.fetchone()

if not user:
    print(f'Erro: Usuário {TENANT_ID} não encontrado!')
    exit(1)

print(f'Usuario encontrado: {user[1]} ({user[2]})\n')

# ============================================================================
# PASSO 1: Limpar dados antigos
# ============================================================================

print('Limpando dados antigos...')
cur.execute('DELETE FROM session_payments WHERE user_id = %s', (TENANT_ID,))
cur.execute('DELETE FROM therapy_sessions WHERE user_id = %s', (TENANT_ID,))
cur.execute('DELETE FROM patients WHERE user_id = %s', (TENANT_ID,))
conn.commit()
print('Dados antigos removidos\n')

# ============================================================================
# PASSO 2: Ler aba "Financ Cobranças" (PACIENTES + SESSÕES)
# ============================================================================

print('PASSO 2: Lendo aba "Financ Cobrancas"...')
ws_cobrancas = wb[wb.sheetnames[0]]

# Estrutura:
# Row 4: nomes dos pacientes
# Rows 5-13: metadata do paciente
# Row 14: headers Pr/Cb
# Rows 15+: pacotes e sessões

pacientes_data = {}

# Encontrar todas as colunas com pacientes (row 4)
for col in range(1, ws_cobrancas.max_column + 1):
    cell = ws_cobrancas.cell(row=4, column=col)
    if cell.value:
        nome = str(cell.value).strip()

        if is_paciente(nome):
            nome_norm = normalizar_nome(nome)
            print(f'  Encontrado paciente: {nome_norm} (col {col})')

            # Ler metadata do paciente
            dia_atendimento = ws_cobrancas.cell(row=5, column=col).value
            horario = ws_cobrancas.cell(row=6, column=col).value
            frequencia = ws_cobrancas.cell(row=7, column=col).value
            data_nascimento_raw = ws_cobrancas.cell(row=8, column=col).value
            data_inicio_raw = ws_cobrancas.cell(row=9, column=col).value
            valor_sessao_raw = ws_cobrancas.cell(row=11, column=col).value
            tipo_fechamento = ws_cobrancas.cell(row=12, column=col).value

            data_nascimento = excel_date_to_python(data_nascimento_raw)
            data_inicio = excel_date_to_python(data_inicio_raw)
            valor_sessao = extract_valor_sessao(valor_sessao_raw)

            pacientes_data[nome_norm] = {
                'column': col,
                'dia_atendimento': str(dia_atendimento) if dia_atendimento else None,
                'horario': str(horario) if horario else None,
                'frequencia': str(frequencia) if frequencia else '1x semana',
                'data_nascimento': data_nascimento,
                'data_inicio': data_inicio or date.today(),
                'valor_sessao': valor_sessao,
                'tipo_fechamento': str(tipo_fechamento) if tipo_fechamento else 'avulso',
                'sessoes': []
            }

            # Ler sessões (rows 19+, colunas col+0=presença, col+1=cobrança, col+2=data)
            for row in range(19, ws_cobrancas.max_row + 1):
                presenca_cell = ws_cobrancas.cell(row=row, column=col).value
                cobranca_cell = ws_cobrancas.cell(row=row, column=col+1).value
                data_cell = ws_cobrancas.cell(row=row, column=col+2).value

                # Presença: 1=presente, 0=ausente
                # Cobrança: s=pago, n=não cobra, d=devedor
                if presenca_cell in [0, 1] and cobranca_cell in ['s', 'n', 'd']:
                    data_sessao = excel_date_to_python(data_cell)

                    if data_sessao:
                        pacientes_data[nome_norm]['sessoes'].append({
                            'data': data_sessao,
                            'presenca': presenca_cell,
                            'cobranca': cobranca_cell
                        })

print(f'\nTotal de pacientes encontrados: {len(pacientes_data)}')
total_sessoes = sum(len(p['sessoes']) for p in pacientes_data.values())
print(f'Total de sessoes encontradas: {total_sessoes}\n')

# ============================================================================
# PASSO 3: Importar pacientes no banco
# ============================================================================

print('PASSO 3: Criando pacientes no banco...')
pacientes_map = {}  # nome -> id

for nome in sorted(pacientes_data.keys()):
    data = pacientes_data[nome]
    patient_id = str(uuid.uuid4())

    # Montar notes com dia/horário
    notes = None
    if data['dia_atendimento'] and data['horario']:
        notes = f"{data['dia_atendimento']} as {data['horario']}"

    # Inserir paciente
    cur.execute('''
        INSERT INTO patients (
            id, user_id, name, patient_status, session_fee, frequency,
            contract_type, started_at, birth_date, notes, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        patient_id,
        TENANT_ID,
        nome,
        'ativo',
        str(data['valor_sessao']),
        data['frequencia'],
        'avulso',  # contract_type enum: 'pacote' or 'avulso'
        data['data_inicio'],
        data['data_nascimento'],
        notes,
        datetime.now()
    ))

    pacientes_map[nome] = patient_id
    print(f'  {nome} ({patient_id})')

conn.commit()
print(f'\n{len(pacientes_map)} pacientes criados\n')

# ============================================================================
# PASSO 4: Importar sessões no banco
# ============================================================================

print('PASSO 4: Criando sessoes no banco...')
sessions_count = 0

for nome, data in pacientes_data.items():
    patient_id = pacientes_map.get(nome)

    if not patient_id:
        continue

    for sessao in data['sessoes']:
        session_id = str(uuid.uuid4())

        # Status da sessão baseado em presença
        # presenca: 1=realizada, 0=nao_realizada
        status = 'realizada' if sessao['presenca'] == 1 else 'nao_realizada'

        # Cobrável baseado em cobrança
        # s=cobrado (chargeable=true), n=não cobra (false), d=devedor (true)
        chargeable = sessao['cobranca'] in ['s', 'd']

        cur.execute('''
            INSERT INTO therapy_sessions (
                id, user_id, patient_id, date, duration, fee,
                status, chargeable, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            session_id,
            TENANT_ID,
            patient_id,
            datetime.combine(sessao['data'], datetime.min.time()),  # date -> datetime
            50,  # duração padrão
            str(data['valor_sessao']),
            status,
            chargeable,
            datetime.now()
        ))

        sessions_count += 1

conn.commit()
print(f'\n{sessions_count} sessoes criadas\n')

# ============================================================================
# PASSO 5: Ler aba "Financ entradas" (PAGAMENTOS)
# ============================================================================

print('PASSO 5: Lendo aba "Financ entradas"...')
ws_entradas = wb[wb.sheetnames[1]]

pagamentos_lista = []

# Estrutura: cada 4 colunas = 1 mês
# Col 1: Data, Cols 2-4: Nome + Valor

for col_offset in range(0, ws_entradas.max_column, 4):
    col = col_offset + 1  # openpyxl é 1-indexed

    # Primeira célula do grupo = data do mês
    data_cell = ws_entradas.cell(row=1, column=col).value

    if data_cell and isinstance(data_cell, datetime):
        data_mes = data_cell.date()
        print(f'  Processando mes: {data_mes.strftime("%Y-%m")} (col {col})')

        # Percorrer linhas deste mês
        for row in range(2, ws_entradas.max_row + 1):
            nome_cell = ws_entradas.cell(row=row, column=col + 1).value
            valor_cell = ws_entradas.cell(row=row, column=col + 2).value

            if nome_cell and valor_cell and isinstance(valor_cell, (int, float)):
                nome = normalizar_nome(str(nome_cell))
                valor = float(valor_cell)

                if is_paciente(nome) and valor > 0:
                    pagamentos_lista.append({
                        'data': data_mes,
                        'paciente': nome,
                        'valor': valor
                    })

print(f'\nTotal de pagamentos encontrados: {len(pagamentos_lista)}\n')

# ============================================================================
# PASSO 6: Importar pagamentos no banco
# ============================================================================

print('PASSO 6: Criando pagamentos no banco...')
payments_count = 0

for pag in pagamentos_lista:
    patient_id = pacientes_map.get(pag['paciente'])

    if patient_id:
        payment_id = str(uuid.uuid4())

        cur.execute('''
            INSERT INTO session_payments (
                id, user_id, patient_id, date, amount,
                method, status, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            payment_id,
            TENANT_ID,
            patient_id,
            pag['data'],
            str(pag['valor']),
            'pix',  # method (enum: pix, card, cash, transfer)
            'paid',  # status (enum: paid, pending, overdue)
            datetime.now()
        ))

        payments_count += 1

conn.commit()
print(f'\n{payments_count} pagamentos criados\n')

# ============================================================================
# Resumo final
# ============================================================================

print('=' * 80)
print('IMPORTACAO CONCLUIDA!')
print('=' * 80)
print(f'\nResumo:')
print(f'  - Pacientes: {len(pacientes_map)}')
print(f'  - Sessoes: {sessions_count}')
print(f'  - Pagamentos: {payments_count}')
print('')

# Fechar conexão
cur.close()
conn.close()

print('Conexao fechada. Importacao finalizada com sucesso!')
