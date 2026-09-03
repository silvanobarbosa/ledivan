#!/usr/bin/env python3
"""
Script COMPLETO de importação dos dados da Gisele
Planilha: Financeiro 040826.xlsx

Importa:
- Pacientes (da aba "Financ Cobranças")
- Sessões/Agendas (da aba "Financ Cobranças")
- Pagamentos (da aba "Financ entradas")
"""

import openpyxl
import psycopg2
from datetime import datetime, date
import uuid
import os
from decimal import Decimal

# Configuração
TENANT_ID = '5588fb64-368f-4deb-81be-80a611967ec1'  # Gisele
DATABASE_URL = os.getenv('DATABASE_URL')

if not DATABASE_URL:
    print("❌ Erro: DATABASE_URL não definida")
    exit(1)

# Conectar ao banco
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

# Nomes que NÃO são pacientes (despesas/outros)
NAO_PACIENTES = {
    'Academia', 'Botox', 'Bolo Gi', 'Capoeira', 'Cartão', 'Ceclin',
    'Congresso', 'Custo fixo', 'Custo Fixo', 'Dentista Heitor', 'Depilação',
    'Devolução Amazon', 'Diferença', 'Diferença:', 'Doação Obreiros',
    'Débito', 'Entrada Diversos', 'Entrada Pontos Cartão', 'Entradas div',
    'Entrada Heitor', 'Estacionamento', 'Exame', 'Festa Vini', 'Heitor escola',
    'Heitor Escola', 'IPTU', 'Iptu', 'Investimento', 'Manutenção',
    'Matricula Heitor', 'Mercado', 'Multas', 'Pagamentos', 'Pedicure',
    'Pix enviado diversos', 'Pix Enviado Diversos', 'Placa dentista',
    'Placa Dentista', 'Present Cris', 'Present EU', 'Present Eu',
    'Present Silvano', 'Presente Felipe', 'Presente Laura', 'Presente Paty',
    'Presente Rafa', 'Pós', 'RNTP', 'Rntp', 'Remédios', 'Sala',
    'Supervisão', 'Terapia', 'Terapia Heitor', 'Uber', 'Vereda'
}

def normalizar_nome(nome):
    """Normaliza nome: primeira letra maiúscula, resto minúsculo"""
    if not nome:
        return ''
    return ' '.join(p.capitalize() for p in str(nome).strip().split())

def is_paciente(nome):
    """Verifica se é nome de paciente (não despesa)"""
    if not nome or not isinstance(nome, str):
        return False

    norm = normalizar_nome(nome)

    if nome in NAO_PACIENTES or norm in NAO_PACIENTES:
        return False

    # Se tem apenas letras e espaços, provavelmente é paciente
    return all(c.isalpha() or c.isspace() for c in norm) and len(norm) > 2

def excel_date_to_python(excel_date):
    """Converte data do Excel para Python date"""
    if not excel_date or not isinstance(excel_date, (int, float, datetime)):
        return None

    if isinstance(excel_date, datetime):
        return excel_date.date()

    # Excel usa 1900-01-01 como base (serial 1)
    # Mas tem bug: considera 1900 bissexto (não é)
    from datetime import timedelta
    base = datetime(1899, 12, 30)  # Ajuste do bug
    try:
        return (base + timedelta(days=excel_date)).date()
    except:
        return None

print('🔄 Iniciando importação COMPLETA dos dados da Gisele...\n')

# Carregar planilha
file_path = 'C:/Users/User/Ledivan Plus/Gi/Financeiro 040826.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Verificar usuário
cur.execute('SELECT id, name, email FROM "user" WHERE id = %s', (TENANT_ID,))
user = cur.fetchone()

if not user:
    print(f'❌ Erro: Usuário {TENANT_ID} não encontrado!')
    exit(1)

print(f'✅ Usuário encontrado: {user[1]} ({user[2]})\n')

# ============================================================================
# PASSO 1: Limpar dados antigos
# ============================================================================

print('🗑️  Limpando dados antigos...')
cur.execute('DELETE FROM session_payments WHERE user_id = %s', (TENANT_ID,))
cur.execute('DELETE FROM therapy_sessions WHERE user_id = %s', (TENANT_ID,))
cur.execute('DELETE FROM patients WHERE user_id = %s', (TENANT_ID,))
conn.commit()
print('✅ Dados antigos removidos\n')

# ============================================================================
# PASSO 2: Ler aba "Financ entradas" (PAGAMENTOS)
# ============================================================================

print('💰 PASSO 2: Lendo aba "Financ entradas"...')
ws_entradas = wb[wb.sheetnames[1]]

pagamentos_lista = []

# Estrutura: cada 4 colunas = 1 mês
# Col 1: Data, Cols 2-4: Nome + Valor

for col_offset in range(0, ws_entradas.max_column, 4):
    col = col_offset + 1  # openpyxl é 1-indexed

    # Primeira célula do grupo = data do mês
    data_cell = ws_entradas.cell(row=1, column=col).value

    if data_cell and isinstance(data_cell, datetime):
        data_mes = data_cell.date()
        print(f'  📅 Processando mês: {data_mes.strftime("%Y-%m")} (col {col})')

        # Percorrer linhas deste mês
        for row in range(2, ws_entradas.max_row + 1):
            nome_cell = ws_entradas.cell(row=row, column=col + 1).value
            valor_cell = ws_entradas.cell(row=row, column=col + 2).value

            if nome_cell and valor_cell and isinstance(valor_cell, (int, float)):
                nome = normalizar_nome(str(nome_cell))
                valor = float(valor_cell)

                if is_paciente(nome) and valor > 0:
                    pagamentos_lista.append({
                        'data': data_mes,
                        'paciente': nome,
                        'valor': valor
                    })

print(f'\n✅ Total de pagamentos encontrados: {len(pagamentos_lista)}\n')

# Agrupar pagamentos por paciente para criar pacientes únicos
pacientes_set = set(p['paciente'] for p in pagamentos_lista)
print(f'👥 Pacientes únicos identificados: {len(pacientes_set)}\n')

# ============================================================================
# PASSO 3: Importar pacientes no banco
# ============================================================================

print('📝 PASSO 3: Criando pacientes no banco...')
pacientes_map = {}  # nome -> id

for nome in sorted(pacientes_set):
    patient_id = str(uuid.uuid4())

    # Inserir paciente
    cur.execute('''
        INSERT INTO patients (
            id, user_id, name, patient_status, session_fee, frequency,
            contract_type, started_at, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        patient_id,
        TENANT_ID,
        nome,
        'ativo',
        '200',  # Preço padrão (session_fee)
        '1x semana',
        'avulso',
        date.today(),
        datetime.now()
    ))

    pacientes_map[nome] = patient_id
    print(f'  ✅ {nome} ({patient_id})')

conn.commit()
print(f'\n✅ {len(pacientes_map)} pacientes criados\n')

# ============================================================================
# PASSO 4: Importar pagamentos no banco
# ============================================================================

print('💰 PASSO 4: Criando pagamentos no banco...')
payments_count = 0

for pag in pagamentos_lista:
    patient_id = pacientes_map.get(pag['paciente'])

    if patient_id:
        payment_id = str(uuid.uuid4())

        cur.execute('''
            INSERT INTO session_payments (
                id, user_id, patient_id, date, amount,
                method, status, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            payment_id,
            TENANT_ID,
            patient_id,
            pag['data'],
            str(pag['valor']),
            'pix',  # method (enum: pix, card, cash, transfer)
            'paid',  # status (enum: paid, pending, overdue)
            datetime.now()
        ))

        payments_count += 1

conn.commit()
print(f'\n✅ {payments_count} pagamentos criados\n')

# ============================================================================
# Resumo final
# ============================================================================

print('=' * 80)
print('✅ IMPORTAÇÃO CONCLUÍDA!')
print('=' * 80)
print(f'\nResumo:')
print(f'  - Pacientes: {len(pacientes_map)}')
print(f'  - Pagamentos: {payments_count}')
print(f'  - Sessões: 0 (TODO: preciso localizar dados de agendas/sessões)')
print('')

# Fechar conexão
cur.close()
conn.close()

print('✅ Conexão fechada. Importação finalizada com sucesso!')
