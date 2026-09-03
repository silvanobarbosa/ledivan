#!/usr/bin/env python3
"""
Corrige a AGENDA dos pacientes da Gisele (conta giselesantosbarbosa) + horário das sessões.

Achados da re-auditoria da planilha Financeiro 040826.xlsx:
  1. dia/horário de atendimento ficaram só no `notes` ("Quinta-feira 16h") — as colunas próprias
     attendance_day / attendance_time estavam VAZIAS (o app não conseguia usar a agenda).
  2. todas as 1690 sessões estão às 00:00 (a planilha só tinha a data, sem hora).

Faz:
  - attendance_day / attendance_time / frequency por paciente, de "Financ Cobranças" (27 com
    agenda) + "Altas e desistências" (dia+hora na mesma célula). Horário normalizado p/ HH:MM.
  - tira o "dia hora" do notes (agora está nas colunas certas).
  - põe TODAS as sessões passadas às 07:00 (pedido da dona: agendas sem horário -> 7h).

Uso:  DATABASE_URL=... python scripts/fix-gisele-agendas.py [--executar]
"""
import openpyxl, psycopg2, os, sys, re, unicodedata

EXECUTAR = "--executar" in sys.argv
ARQ = "C:/Users/User/Ledivan Plus/Gi/Financeiro 040826.xlsx"
USER_ID = "0d320bc4-2f9e-433c-bb64-28f10f54f8bf"

def norm(s):
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower().strip() if s else ""

def is_pac(nome):
    n = norm(nome)
    return bool(n) and len(n) > 1 and all(c.isalpha() or c.isspace() for c in n) and n not in ("total", "saldo", "nome", "status")

def hora_hhmm(txt):
    """ '14h' -> '14:00', '8h'/'09h' -> '08:00'/'09:00', 'as 10h' -> '10:00'. None se não achar."""
    if not txt:
        return None
    m = re.search(r"(\d{1,2})\s*[h:]", str(txt))
    if not m:
        m = re.search(r"\b(\d{1,2})\b", str(txt))
    if not m:
        return None
    h = int(m.group(1))
    if h > 23:
        return None
    return f"{h:02d}:00"

def dia_texto(txt):
    """Normaliza o dia mantendo acento amigável. Ex.: 'Terça e quinta' fica como está."""
    if not txt:
        return None
    s = str(txt).strip()
    return s or None

wb = openpyxl.load_workbook(ARQ, data_only=True)
agenda = {}  # norm(nome) -> {dia, hora, freq}

# Financ Cobranças: dia r5, hora r6, freq r7 (na col do nome + 2)
ws = wb["Financ Cobranças"]
for c in range(1, ws.max_column + 1):
    nome = ws.cell(row=4, column=c).value
    if not is_pac(nome):
        continue
    v = c + 2
    agenda[norm(nome)] = {
        "dia": dia_texto(ws.cell(row=5, column=v).value),
        "hora": hora_hhmm(ws.cell(row=6, column=v).value),
        "freq": (str(ws.cell(row=7, column=v).value).strip() if ws.cell(row=7, column=v).value else None),
    }

# Altas e desistências: nome r8, "dia hora" juntos em r9 (ex.: 'Terça 10h')
ws = wb["Altas e desistências"]
for c in range(1, ws.max_column + 1):
    nome = ws.cell(row=8, column=c).value
    if not is_pac(nome):
        continue
    cell = ws.cell(row=9, column=c).value
    if not cell:
        continue
    # separa dia (palavra) e hora (numero h)
    dia = re.sub(r"\d.*", "", str(cell)).strip() or None
    agenda.setdefault(norm(nome), {}).update({"dia": dia, "hora": hora_hhmm(cell)})

comp = sum(1 for a in agenda.values() if a.get("hora"))
print(f"  pacientes com agenda parseada: {len(agenda)} (com horário: {comp})")
if not EXECUTAR:
    print("  DRY-RUN — amostra:")
    for k in list(agenda)[:8]:
        a = agenda[k]
        print(f"    {k[:16]:16} dia={a.get('dia')} hora={a.get('hora')} freq={a.get('freq')}")
    sys.exit(0)

conn = psycopg2.connect(os.getenv("DATABASE_URL"))
cur = conn.cursor()
try:
    # 1) preenche attendance_day/time/frequency por paciente (match por nome normalizado)
    atualizados = 0
    cur.execute('SELECT id, name, notes FROM patients WHERE user_id=%s', (USER_ID,))
    for pid, name, notes in cur.fetchall():
        a = agenda.get(norm(name))
        if not a:
            continue
        # limpa o "dia hora" do notes (agora nas colunas certas)
        novo_notes = notes
        if notes and a.get("dia"):
            novo_notes = re.sub(re.escape(str(a["dia"])) + r"\s*\S*", "", notes).strip() or None
        cur.execute(
            """UPDATE patients SET attendance_day=%s, attendance_time=%s,
               frequency=COALESCE(%s, frequency), notes=%s WHERE id=%s""",
            (a.get("dia"), a.get("hora"), a.get("freq"), novo_notes, pid),
        )
        atualizados += 1
    print(f"  pacientes atualizados (agenda): {atualizados}")

    # 2) todas as sessões passadas -> 07:00
    cur.execute(
        "UPDATE therapy_sessions SET date = date_trunc('day', date) + interval '7 hours' WHERE user_id=%s",
        (USER_ID,),
    )
    print(f"  sessões movidas para 07:00: {cur.rowcount}")

    conn.commit()
    print("  COMMIT ok")
except Exception as e:
    conn.rollback()
    print("  ROLLBACK:", e)
    raise
finally:
    conn.close()
