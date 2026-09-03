#!/usr/bin/env python3
"""
Reimportação COMPLETA e validada dos dados da Gisele (conta correta giselesantosbarbosa).

Fonte: Gi/Financeiro 040826.xlsx (04/08/2026). Planilha irregular, dados em 6 abas.
Alvo:  user_id 0d320bc4-... (giselesantosbarbosa@gmail.com) — a conta que a dona confirmou.

Por que v3: o v2 lia só a aba "Financ Cobranças" (27 pacientes) e tinha offset na metadata; o
import de junho ficou com buraco em 2026 (jan/fev faltando) e um "2026-06" fantasma. Aqui a
verdade vem de TODAS as abas, com reconciliação no fim (planilha == banco).

Regras:
- Paciente = qualquer nome de pessoa em Financ Cobranças (row4), Estatística (col Nome) OU
  Financ entradas (nome de pagamento). Despesas conhecidas (NAO_PACIENTES) ficam de fora.
- Metadata de Financ Cobranças fica na coluna do nome + 2 (offset real conferido no arquivo).
- Todo pagamento tem que casar com um paciente; se o nome só aparece em pagamentos, o paciente
  é criado. Meta: ZERO pagamento órfão.

Uso:
  DATABASE_URL=... python scripts/import-gisele-v3.py            # DRY-RUN (só reporta)
  DATABASE_URL=... python scripts/import-gisele-v3.py --executar # escreve no banco
"""
import openpyxl, psycopg2, os, sys, uuid, re, unicodedata
from datetime import datetime, date, timedelta
from collections import defaultdict

EXECUTAR = "--executar" in sys.argv
ARQ = "C:/Users/User/Ledivan Plus/Gi/Financeiro 040826.xlsx"
USER_ID = "0d320bc4-2f9e-433c-bb64-28f10f54f8bf"  # giselesantosbarbosa@gmail.com

NAO_PACIENTES = {
    "academia","botox","capoeira","cartao","ceclin","congresso","custo fixo","dentista heitor",
    "depilacao","devolucao amazon","diferenca","doacao obreiros","debito","entrada diversos",
    "entrada pontos cartao","entradas div","entrada heitor","estacionamento","exame","festa vini",
    "heitor escola","iptu","investimento","manutencao","matricula heitor","mercado","multas",
    "pagamentos","pedicure","pix enviado diversos","placa dentista","present cris","present eu",
    "present silvano","presente felipe","presente laura","presente paty","presente rafa","pos",
    "rntp","remedios","sala","supervisao","terapia","terapia heitor","uber","vereda","total",
    "saldo","status","nome","bolo gi","diferenca:",
}

def norm(s):
    if s is None: return ""
    return unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower().strip()

def titulo(s):
    return " ".join(p.capitalize() for p in str(s).strip().split())

def is_pac(nome):
    n = norm(nome)
    if not n or n in NAO_PACIENTES or len(n) < 2: return False
    return all(c.isalpha() or c.isspace() for c in n)

def to_dt(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, date): return datetime.combine(v, datetime.min.time())
    if isinstance(v, str):
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", v)
        if m:
            d, mo, y = map(int, m.groups()); return datetime(y, mo, d)
        return None
    if isinstance(v, (int, float)):
        try: return datetime(1899,12,30) + timedelta(days=v)
        except: return None
    return None

def valor_num(txt, default=None):
    if txt is None: return default
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(txt).replace(",", "."))
    return float(m.group(1)) if m else default

wb = openpyxl.load_workbook(ARQ, data_only=True)

# nome_norm -> dados do paciente
pac = {}
def get_pac(nome):
    k = norm(nome)
    if k not in pac:
        pac[k] = {"nome": titulo(nome), "status": None, "phone": None, "birth": None,
                  "started": None, "fee": None, "freq": None, "contract": None,
                  "dia": None, "hora": None, "sessoes": [],
                  "prospect_date": None, "prospect_fechou": None}
    return pac[k]

# --- Financ Cobranças: pacientes + metadata (col+2) + sessões ---
ws = wb["Financ Cobranças"]
cols_pac = [(c, ws.cell(row=4, column=c).value) for c in range(1, ws.max_column+1)
            if is_pac(ws.cell(row=4, column=c).value)]
for c, nome in cols_pac:
    p = get_pac(nome)
    v = c + 2  # metadata e sessões: valores na coluna do nome + 2
    p["dia"]   = ws.cell(row=5, column=v).value
    p["hora"]  = ws.cell(row=6, column=v).value
    p["freq"]  = ws.cell(row=7, column=v).value or p["freq"]
    p["birth"] = to_dt(ws.cell(row=8, column=v).value) or p["birth"]
    p["started"] = to_dt(ws.cell(row=9, column=v).value) or p["started"]
    p["fee"]   = valor_num(ws.cell(row=11, column=v).value, p["fee"])
    fech = ws.cell(row=12, column=v).value
    p["contract"] = "pacote" if fech and "pacote" in norm(fech) else "avulso"
    if p["status"] is None: p["status"] = "ativo"
    # Sessões (rows 19+): para o nome na col c, presença fica na PRÓPRIA col c, cobrança em c+1,
    # data em c+2 — conferido no arquivo (ELOÁ col4 -> presença col4, cb col5, data col6).
    for r in range(19, ws.max_row+1):
        pres = ws.cell(row=r, column=c).value
        cob  = ws.cell(row=r, column=c+1).value
        dt   = to_dt(ws.cell(row=r, column=c+2).value)
        if pres in (0, 1) and norm(cob) in ("s","n","d") and dt:
            p["sessoes"].append({"data": dt, "pres": pres, "cob": norm(cob)})

# --- Altas e desistências: pacientes que sairam (nomes row 8, sessões rows 11+) ---
# Mesmo layout horizontal do Cobranças, mas de alta/desistência. Aqui a coluna do nome guarda as
# DATAS das sessões (sem coluna de presença/cobrança separada), com linhas-cabeçalho de pacote no
# meio. Tratamos cada data como sessão realizada e cobrável; ignoramos linhas de texto (Pacote…).
ws = wb["Altas e desistências"]
for c in range(1, ws.max_column+1):
    nome = ws.cell(row=8, column=c).value
    if not is_pac(nome): continue
    p = get_pac(nome)
    if p["status"] is None or p["status"] == "ativo": p["status"] = "inativo"
    hora = ws.cell(row=9, column=c).value
    if hora and not p["hora"]: p["hora"] = hora
    for r in range(11, ws.max_row+1):
        dt = to_dt(ws.cell(row=r, column=c).value)
        # só conta se a célula é REALMENTE uma data (não texto tipo "Pacote 2" nem "17/12 - não fechou")
        raw = ws.cell(row=r, column=c).value
        if dt and isinstance(raw, (datetime, date)):
            p["sessoes"].append({"data": dt, "pres": 1, "cob": "s"})

# --- Estatística: status, phone, birth, started ---
ws = wb["Estatística"]
for r in range(3, ws.max_row+1):
    nome = ws.cell(row=r, column=3).value
    if not is_pac(nome): continue
    p = get_pac(nome)
    st = norm(ws.cell(row=r, column=2).value)
    if st in ("ativo",): p["status"] = "ativo"
    elif st in ("parou","desistiu","inativo","alta"): p["status"] = p["status"] or "inativo"
    ph = ws.cell(row=r, column=4).value
    if ph and not p["phone"]: p["phone"] = re.sub(r"\D", "", str(ph)) or None
    if not p["birth"]: p["birth"] = to_dt(ws.cell(row=5, column=r).value) if False else p["birth"]
    b = to_dt(ws.cell(row=r, column=5).value)
    if b and not p["birth"]: p["birth"] = b
    s = to_dt(ws.cell(row=r, column=8).value)
    if s and not p["started"]: p["started"] = s

# --- Prospecção: leads. Vira paciente status='prospect' SE ainda não existe (não rebaixa quem
#     já fechou e virou ativo). Guarda data e se fechou. ---
ws = wb["Prospecção"]
for r in range(2, ws.max_row+1):
    nome = ws.cell(row=r, column=3).value
    if not is_pac(nome): continue
    novo = norm(nome) not in pac
    p = get_pac(nome)
    if novo:
        p["status"] = "prospect"
        p["prospect_date"] = to_dt(ws.cell(row=r, column=2).value)
        p["prospect_fechou"] = ws.cell(row=r, column=4).value

# --- Financ entradas: pagamentos (17 meses) ---
ws = wb["Financ entradas"]
meses = [(c, ws.cell(row=1, column=c).value) for c in range(1, ws.max_column+1)
         if isinstance(ws.cell(row=1, column=c).value, (datetime, date))]
pagamentos = []
for c, mes in meses:
    mdt = to_dt(mes)
    for r in range(2, ws.max_row+1):
        nome = ws.cell(row=r, column=c+1).value
        val  = ws.cell(row=r, column=c+2).value
        if nome and isinstance(val, (int, float)) and val > 0 and is_pac(nome):
            get_pac(nome)  # garante que o paciente existe (cria se só aparece aqui)
            pagamentos.append({"nome": norm(nome), "data": mdt, "valor": float(val)})

# defaults para pacientes que só apareceram em pagamentos
for k, p in pac.items():
    if p["status"] is None: p["status"] = "ativo"
    if p["fee"] is None: p["fee"] = 200.0
    if p["contract"] is None: p["contract"] = "avulso"
    if p["started"] is None: p["started"] = datetime(2025,1,1)

total_sess = sum(len(p["sessoes"]) for p in pac.values())
total_val = sum(x["valor"] for x in pagamentos)

print("="*70)
print(f"  DRY-RUN — nada escrito.  (rode com --executar para gravar)" if not EXECUTAR else "  EXECUTANDO — vai gravar no banco")
print("="*70)
print(f"  Pacientes (union de todas as abas): {len(pac)}")
print(f"  Sessões (Financ Cobranças):         {total_sess}")
print(f"  Pagamentos (Financ entradas):       {len(pagamentos)} = R$ {total_val:.2f}")
orfaos = [x for x in pagamentos if x["nome"] not in pac]
print(f"  Pagamentos órfãos (sem paciente):   {len(orfaos)}  (tem que ser 0)")
com_fee = sum(1 for p in pac.values() if p['fee'])
com_birth = sum(1 for p in pac.values() if p['birth'])
print(f"  Pacientes com valor/sessão:  {com_fee}/{len(pac)}   com nascimento: {com_birth}/{len(pac)}")

if not EXECUTAR:
    print("\n  amostra de 5 pacientes:")
    for k in sorted(pac)[:5]:
        p = pac[k]
        print(f"    {p['nome']:20} status={p['status']:7} fee={p['fee']} nasc={str(p['birth'])[:10]} sessoes={len(p['sessoes'])}")
    sys.exit(0)

# ====================== ESCRITA (transação única) ======================
url = os.getenv("DATABASE_URL")
conn = psycopg2.connect(url); cur = conn.cursor()
try:
    cur.execute('SELECT email FROM "user" WHERE id=%s', (USER_ID,))
    u = cur.fetchone()
    if not u: raise SystemExit(f"usuário {USER_ID} não existe")
    print(f"\n  alvo: {u[0]}")

    cur.execute("DELETE FROM session_payments WHERE user_id=%s", (USER_ID,))
    cur.execute("DELETE FROM therapy_sessions WHERE user_id=%s", (USER_ID,))
    cur.execute("DELETE FROM patients WHERE user_id=%s", (USER_ID,))

    pid = {}
    for k, p in pac.items():
        i = str(uuid.uuid4()); pid[k] = i
        notes = f"{p['dia']} {p['hora']}".strip() if (p['dia'] or p['hora']) else None
        cur.execute("""INSERT INTO patients
            (id,user_id,name,phone,session_fee,frequency,patient_status,contract_type,
             started_at,birth_date,notes,prospect_date,prospect_fechou,created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())""",
            (i, USER_ID, p["nome"], p["phone"], str(p["fee"]), (str(p["freq"]) if p["freq"] else None),
             p["status"], p["contract"], p["started"], p["birth"], notes,
             p["prospect_date"], (str(p["prospect_fechou"]) if p["prospect_fechou"] else None)))

    ns = 0
    for k, p in pac.items():
        for s in p["sessoes"]:
            status = "realizada" if s["pres"] == 1 else "nao_realizada"
            chargeable = s["cob"] in ("s","d")
            cur.execute("""INSERT INTO therapy_sessions
                (id,user_id,patient_id,date,duration,fee,status,chargeable,created_at)
                VALUES (%s,%s,%s,%s,50,%s,%s,%s,now())""",
                (str(uuid.uuid4()), USER_ID, pid[k], s["data"], str(p["fee"]), status, chargeable))
            ns += 1

    npay = 0
    for x in pagamentos:
        if x["nome"] not in pid: continue
        cur.execute("""INSERT INTO session_payments
            (id,user_id,patient_id,date,amount,method,status,created_at)
            VALUES (%s,%s,%s,%s,%s,'pix','paid',now())""",
            (str(uuid.uuid4()), USER_ID, pid[x["nome"]], x["data"], str(x["valor"])))
        npay += 1

    # reconciliação DENTRO da transação, antes do commit
    cur.execute("SELECT count(*) FROM patients WHERE user_id=%s", (USER_ID,)); dbp = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM therapy_sessions WHERE user_id=%s", (USER_ID,)); dbs = cur.fetchone()[0]
    cur.execute("SELECT count(*), coalesce(sum(amount),0) FROM session_payments WHERE user_id=%s", (USER_ID,)); dbpay = cur.fetchone()
    print(f"  inseridos: pacientes={dbp} sessoes={dbs} pagamentos={dbpay[0]} R${dbpay[1]}")
    ok = (dbp == len(pac) and dbs == ns and dbpay[0] == npay)
    if not ok:
        raise SystemExit(f"RECONCILIAÇÃO FALHOU (parse: p={len(pac)} s={ns} pay={npay}) — rollback")
    conn.commit()
    print("  COMMIT ok — reconciliação bateu.")
except Exception as e:
    conn.rollback()
    print(f"  ROLLBACK: {e}")
    raise
finally:
    conn.close()
